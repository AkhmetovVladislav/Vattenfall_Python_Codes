from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from copy import copy
import datetime
import os
import shutil
import time

def ausführen():
    try:

        # Benutzernamen abrufen
        benutzername = os.getlogin()

        # Quellverzeichnis
        quelle_verzeichnis = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcing - Dokumente\Fahrplanmanagement\APM_RLM_ZW_TABELLEN'

        # Zielverzeichnis
        ziel_verzeichnis = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcing - Dokumente\IT\Tools\Vladi\Excel-Ausfüller\source\VES_APM_RLM_SLP'

        # Liste aller Dateien im Quellverzeichnis
        dateien = os.listdir(quelle_verzeichnis)

        # Filtern nach Dateien, die "APM_RLM_SLP" im Namen enthalten
        passende_dateien = [datei for datei in dateien if "APM_RLM_SLP" in datei]

        # Die neueste Datei auswählen (basiert auf dem Dateinamen, nicht auf dem Änderungsdatum)
        neueste_datei = passende_dateien[-1]

        # Den vollständigen Pfad zur neuesten Datei erstellen
        quelle_pfad = os.path.join(quelle_verzeichnis, neueste_datei)

        # Ziel-Pfad für die kopierte Datei
        ziel_pfad = os.path.join(ziel_verzeichnis, "APM_RLM_SLP.xlsx")

        # Die neueste Datei kopieren und umbenennen
        shutil.copy(quelle_pfad, ziel_pfad)

        # Zieldatei "VES_APM_RLM_SLP.xlsx" mit dem betrachteten Reiter "RLM"

        zieldatei_pfad = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcingPowerBI - Dokumente\IT\PowerBI_Masterdateien\Excel_Masterdateien\VES_APM_RLM_SLP.xlsx'
        Zieldatei = load_workbook(zieldatei_pfad)
        Zieldatei_Reiter = Zieldatei["RLM"]

        # Quelldatei "APM_RLM_SLP.xlsx" mit dem betrachteten Reiter "RLM"

        quelldatei_pfad = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcing - Dokumente\IT\Tools\Vladi\Excel-Ausfüller\source\VES_APM_RLM_SLP\APM_RLM_SLP.xlsx'
        Quelldatei = load_workbook(quelldatei_pfad)
        Quelldatei_Reiter = Quelldatei["RLM"]

        # Hilfstabelle "VNB_Bilanzierungsgebiete_EIC_01-2024.xlsx" mit dem betrachteten Reiter "ALLE", wo alle Bilanzierungsgebiete und Stromnetzbetreiber aufgelistet sind

        hilfstabelle_pfad = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcing - Dokumente\IT\Tools\Vorlagen_Übersetzer_Analysen\VNB_Bilanzierungsgebiete_EIC_01-2024.xlsx'
        Hilfstabelle = load_workbook(hilfstabelle_pfad)
        Hilfstabelle_Reiter = Hilfstabelle["ALLE"]

        # zusätzliche Hilfstabelle "hilfstabelle.xlsx" mit dem betrachteten Reiter "Hilfstabelle", wo alle Bilanzierungsgebiete und Stromnetzbetreiber aufgelistet sind

        zusaetzliche_hilfstabelle_pfad = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcing - Dokumente\IT\Tools\Vladi\Excel-Ausfüller\source\VES_APM_RLM_SLP\hilfstabelle.xlsx'
        zusaetzliche_Hilfstabelle = load_workbook(zusaetzliche_hilfstabelle_pfad)
        zusaetzliche_Hilfstabelle_Reiter = zusaetzliche_Hilfstabelle["Hilfstabelle"]

        # Bereinige die Zieldatei
        Zieldatei_Reiter.delete_cols(1,23)

        # Definiere die Überschriften
        headers = ["EXPORTIERT_AM", "IMPORTIERT_AM", "REGELZONE_EIC", "BILANZIERUNGSGEBIET_EIC", "VERTRAG", "MALO", "JVP_ISU", "JVP_ECOUNT", "JVP_RESULT", "HOECHSTLEISTUNG_RLM", "BENH", "EINZUGSDATUM", "AUSZUGSDATUM", "PROFILBEZ_VNB", "ZEITREIHENTYP", "GUELTIG_AB", "TARIF_ISU", "MESSLOKATION", "LASTPROFIL_APM", "RLM_LITE", "PLZ", "VNB_NAME"]

        # Schreibe die Überschriften in die Zieldatei
        for col_num, header in enumerate(headers, start=1):
            Zieldatei_Reiter.cell(row=1, column=col_num, value=header)

        # Finde das Ende der Daten in Spalte D der Quelldatei
        end_row = Quelldatei_Reiter.max_row

        # Kopiere und füge die Werte von Spalte N in Spalte A ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=14).value
            # Überprüfe, ob der Wert ein Datum ist
            if isinstance(value, datetime.datetime):
                # Konvertiere das Datum in das richtige Format, falls erforderlich
                value = value.strftime("%d.%m.%Y")
            # Füge den Wert in die entsprechende Zelle der Spalte A der Zieldatei ein
            Zieldatei_Reiter.cell(row=row, column=1).value = value

        # Kopiere und füge die Werte von Spalte O in Spalte B ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=15).value
            # Überprüfe, ob der Wert ein Datum ist
            if isinstance(value, datetime.datetime):
                # Konvertiere das Datum in das richtige Format, falls erforderlich
                value = value.strftime("%d.%m.%Y")
            # Füge den Wert in die entsprechende Zelle der Spalte B der Zieldatei ein
            Zieldatei_Reiter.cell(row=row, column=2).value = value

        # Kopiere und füge die Werte von Spalte B in Spalte C ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=2).value
            Zieldatei_Reiter.cell(row=row, column=3, value=value)

        # Kopiere und füge die Werte von Spalte C in Spalte D ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=3).value
            Zieldatei_Reiter.cell(row=row, column=4, value=value)

        # Kopiere und füge die Werte von Spalte A in Spalte E ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=1).value
            Zieldatei_Reiter.cell(row=row, column=5, value=value)

        # Kopiere und füge die Werte von Spalte R in Spalte F ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=18).value
            Zieldatei_Reiter.cell(row=row, column=6, value=value)

        # Kopiere und füge die Werte von Spalte J in Spalte G und Spalte I ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=10).value
            Zieldatei_Reiter.cell(row=row, column=7, value=value)
            Zieldatei_Reiter.cell(row=row, column=9, value=value)

        # Kopiere und füge die Werte von Spalte K in Spalte J ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=11).value
            Zieldatei_Reiter.cell(row=row, column=10, value=value)

        # Kopiere und füge die Werte von Spalte E in Spalte L ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=5).value
            # Überprüfe, ob der Wert ein Datum ist
            if isinstance(value, datetime.datetime):
                # Konvertiere das Datum in das richtige Format, falls erforderlich
                value = value.strftime("%d.%m.%Y")
            # Füge den Wert in die entsprechende Zelle der Spalte L der Zieldatei ein
            Zieldatei_Reiter.cell(row=row, column=12).value = value

        # Kopiere und füge die Werte von Spalte F in Spalte M ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=6).value
            # Überprüfe, ob der Wert ein Datum ist
            if isinstance(value, datetime.datetime):
                # Konvertiere das Datum in das richtige Format, falls erforderlich
                value = value.strftime("%d.%m.%Y")
            # Füge den Wert in die entsprechende Zelle der Spalte M der Zieldatei ein
            Zieldatei_Reiter.cell(row=row, column=13).value = value

        # Kopiere und füge die Werte von Spalte G in Spalte N ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=7).value
            Zieldatei_Reiter.cell(row=row, column=14, value=value)

        # Kopiere und füge die Werte von Spalte H in Spalte O ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=8).value
            Zieldatei_Reiter.cell(row=row, column=15, value=value)

        # Kopiere und füge die Werte von Spalte I in Spalte P ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=9).value
            # Überprüfe, ob der Wert ein Datum ist
            if isinstance(value, datetime.datetime):
                # Konvertiere das Datum in das richtige Format, falls erforderlich
                value = value.strftime("%d.%m.%Y")
            # Füge den Wert in die entsprechende Zelle der Spalte P der Zieldatei ein
            Zieldatei_Reiter.cell(row=row, column=16).value = value

        # Kopiere und füge die Werte von Spalte L in Spalte Q ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=12).value
            Zieldatei_Reiter.cell(row=row, column=17, value=value)

        # Kopiere und füge die Werte von Spalte S in Spalte T ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=19).value
            Zieldatei_Reiter.cell(row=row, column=20, value=value)

        # Kopiere und füge die Werte von Spalte M in Spalte U ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=13).value
            Zieldatei_Reiter.cell(row=row, column=21, value=value)

        # Die Werte für Spalte K eintragen (Spalte K = Spalte I / Spalte J)
        for row in range(2, end_row + 1):
            # Lies die Werte von Spalte I und J
            value_I = Zieldatei_Reiter.cell(row=row, column=9).value
            value_J = Zieldatei_Reiter.cell(row=row, column=10).value
            # Berechne den Wert für Spalte K
            if value_J is not None and value_J != 0:
                result = round(value_I / value_J)
            else:
                result = None
            # Schreibe den Wert in Spalte K
            Zieldatei_Reiter.cell(row=row, column=11).value = result

        # die Spalte "VNB_NAME" ausfüllen
            
        # Initialisiere eine Variable, um die Anzahl von "kein Wert" zu zählen
        no_value_count = 0

        # Durchlaufe alle Zellen der Spalte D (ab D2 bis zum Ende) der Zieldatei
        for row in range(2, Zieldatei_Reiter.max_row + 1):
            # Lies den Wert der Zelle in Spalte D der Zieldatei
            value_D = Zieldatei_Reiter.cell(row=row, column=4).value
            
            # Initialisiere einen Flag, um festzustellen, ob ein Wert gefunden wurde
            found = False
            
            # Durchlaufe alle Zeilen der Spalte F der Hilfstabelle, um den Wert zu finden
            for hilfs_row in range(3, Hilfstabelle_Reiter.max_row + 1):
                # Lies den Wert der Zelle in Spalte F der Hilfstabelle
                value_F = Hilfstabelle_Reiter.cell(row=hilfs_row, column=6).value
                
                # Überprüfe, ob der Wert übereinstimmt
                if value_F == value_D:
                    # Wenn der Wert übereinstimmt, schreibe den Wert aus Spalte B der Hilfstabelle in Spalte V der Zieldatei
                    Zieldatei_Reiter.cell(row=row, column=22).value = Hilfstabelle_Reiter.cell(row=hilfs_row, column=2).value
                    found = True
                    break
            
            # Wenn kein übereinstimmender Wert in der ersten Hilfstabelle gefunden wurde, suche in der zusätzlichen Hilfstabelle
            if not found:
                # Durchlaufe alle Zeilen der Spalte A der zusätzlichen Hilfstabelle, um den Wert zu finden
                for zus_row in range(2, zusaetzliche_Hilfstabelle_Reiter.max_row + 1):
                    # Lies den Wert der Zelle in Spalte A der zusätzlichen Hilfstabelle
                    zus_value_A = zusaetzliche_Hilfstabelle_Reiter.cell(row=zus_row, column=1).value
                    
                    # Überprüfe, ob der Wert übereinstimmt
                    if zus_value_A == value_D:
                        # Wenn der Wert übereinstimmt, schreibe den Wert aus Spalte B der zusätzlichen Hilfstabelle in Spalte V der Zieldatei
                        Zieldatei_Reiter.cell(row=row, column=22).value = zusaetzliche_Hilfstabelle_Reiter.cell(row=zus_row, column=2).value
                        found = True
                        break

            # Falls kein übereinstimmender Wert gefunden wurde, schreibe "kein Wert" in Spalte V der Zieldatei
            if not found:
                Zieldatei_Reiter.cell(row=row, column=22).value = "kein Wert"
                no_value_count += 1

        # Wenn "kein Wert" geschrieben wurde, gib die entsprechende Meldung aus
        if no_value_count > 0:
            print(f"In der Datei 'VES_APM_RLM_SLP.xlsx' gibt es {no_value_count} fehlende VNB in der Spalte V. Sie haben 'kein Wert' in der Spalte V")

        time.sleep(60)

        # Speichere die Änderungen
        Zieldatei.save(zieldatei_pfad)

        # Die neueste Datei im Quellverzeichnis löschen
        os.remove(quelldatei_pfad)

        print('-------------------------------------------------------------------------------------')
        print('--------die Excel-Datei "VES_APM_RLM_SLP.xlsx" wurde erfolgreich aktualisiert--------')
        print('-------------------------------------------------------------------------------------')

        return True
    
    except Exception as e:

        print(f'Ein Fehler beim Aktualisieren der Datei ""VES_APM_RLM_SLP.xlsx"" ist aufgetreten: {e}')
        return False
    
# Um das Programm auszuführen, soll der Teil unten auskommentiert werden
    
# ausführen()
