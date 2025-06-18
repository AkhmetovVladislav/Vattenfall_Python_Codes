from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from copy import copy
import datetime
import os
import shutil
import time

def ausführen():
    try:

        # Benutzernamen abrufen
        benutzername = os.getlogin()

        # Quellverzeichnis
        quelle_verzeichnis = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcing - Dokumente\Fahrplanmanagement\APM_RLM_ZW_TABELLEN'

        # Zielverzeichnis
        ziel_verzeichnis = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcing - Dokumente\IT\Tools\Vladi\Excel-Ausfüller\source\VES_SLP_Strom'

        # Liste aller Dateien im Quellverzeichnis
        dateien = os.listdir(quelle_verzeichnis)

        # Filtern nach Dateien, die "APM_RLM_SLP" im Namen enthalten
        passende_dateien = [datei for datei in dateien if "APM_RLM_SLP" in datei]

        # Die neueste Datei auswählen (basiert auf dem Dateinamen, nicht auf dem Änderungsdatum)
        neueste_datei = passende_dateien[-1]

        # Den vollständigen Pfad zur neuesten Datei erstellen
        quelle_pfad = os.path.join(quelle_verzeichnis, neueste_datei)

        # Ziel-Pfad für die kopierte Datei
        ziel_pfad = os.path.join(ziel_verzeichnis, "APM_RLM_SLP.xlsx")

        # Die neueste Datei kopieren und umbenennen
        shutil.copy(quelle_pfad, ziel_pfad)

        # Zieldatei "VES_SLP_Strom.xlsx" mit dem betrachteten Reiter "SLP"

        zieldatei_pfad = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcingPowerBI - Dokumente\IT\PowerBI_Masterdateien\Excel_Masterdateien\VES_SLP_Strom.xlsx'
        Zieldatei = load_workbook(zieldatei_pfad)
        Zieldatei_Reiter = Zieldatei["SLP"]

        # Quelldatei "APM_RLM_SLP.xlsx" mit dem betrachteten Reiter "SLP"

        quelldatei_pfad = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcing - Dokumente\IT\Tools\Vladi\Excel-Ausfüller\source\VES_SLP_Strom\APM_RLM_SLP.xlsx'
        Quelldatei = load_workbook(quelldatei_pfad)
        Quelldatei_Reiter = Quelldatei["SLP"]

        # Bereinige die Zieldatei
        Zieldatei_Reiter.delete_cols(1,20)

        # Definiere die Überschriften
        headers = ["KLZ", "FTZ", "GUELTIG_AB", "GUELTIG_BIS", "JAHRESARBEIT", "REGELZONE", "ANZAHL_VS", "LAUF_NR", "RECORDS_EXPECTED", "CREATED", "NETZ_NR_ISU", "PROFILBEZ_VNB", "ZEITREIHENTYP", "BILANZIERUNGSGEBIET_EIC", "BEMERKUNG", "PROFROLE", "Gesellschaft", "EIC"]

        # Schreibe die Überschriften in die Zieldatei
        for col_num, header in enumerate(headers, start=1):
            Zieldatei_Reiter.cell(row=1, column=col_num, value=header)

        # Finde das Ende der Daten in Spalte D der Quelldatei
        end_row = Quelldatei_Reiter.max_row

        # Kopiere und füge die Werte von Spalte D in Spalte C ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=4).value
            # Überprüfe, ob der Wert ein Datum ist
            if isinstance(value, datetime.datetime):
                # Konvertiere das Datum in das richtige Format, falls erforderlich
                value = value.strftime("%d.%m.%Y")
            # Füge den Wert in die entsprechende Zelle der Spalte C der Zieldatei ein
            Zieldatei_Reiter.cell(row=row, column=3).value = value

        # Kopiere und füge die Werte von Spalte E in Spalte D ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=5).value
            # Überprüfe, ob der Wert ein Datum ist
            if isinstance(value, datetime.datetime):
                # Konvertiere das Datum in das richtige Format, falls erforderlich
                value = value.strftime("%d.%m.%Y")
            # Füge den Wert in die entsprechende Zelle der Spalte D der Zieldatei ein
            Zieldatei_Reiter.cell(row=row, column=4).value = value

        # Kopiere und füge die Werte von Spalte F in Spalte E ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=6).value
            Zieldatei_Reiter.cell(row=row, column=5, value=value)

        # Kopiere und füge die Werte von Spalte A in Spalte F ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=1).value
            Zieldatei_Reiter.cell(row=row, column=6, value=value)

        # Kopiere und füge die Werte von Spalte G in Spalte G ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=7).value
            Zieldatei_Reiter.cell(row=row, column=7, value=value)

        # Kopiere und füge die Werte von Spalte K in Spalte H ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=11).value
            Zieldatei_Reiter.cell(row=row, column=8, value=value)

        # Kopiere und füge die Werte von Spalte L in Spalte I ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=12).value
            Zieldatei_Reiter.cell(row=row, column=9, value=value)

        # Kopiere und füge die Werte von Spalte M in Spalte J ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=13).value
            # Überprüfe, ob der Wert ein Datum ist
            if isinstance(value, datetime.datetime):
                # Konvertiere das Datum in das richtige Format, falls erforderlich
                value = value.strftime("%d.%m.%Y")
            # Füge den Wert in die entsprechende Zelle der Spalte J der Zieldatei ein
            Zieldatei_Reiter.cell(row=row, column=10).value = value

        # Kopiere und füge die Werte von Spalte H in Spalte K ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=8).value
            Zieldatei_Reiter.cell(row=row, column=11, value=value)

        # Kopiere und füge die Werte von Spalte C in Spalte L ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=3).value
            Zieldatei_Reiter.cell(row=row, column=12, value=value)

        # Kopiere und füge die Werte von Spalte I in Spalte M ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=9).value
            Zieldatei_Reiter.cell(row=row, column=13, value=value)

        # Kopiere und füge die Werte von Spalte B in Spalte N ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=2).value
            Zieldatei_Reiter.cell(row=row, column=14, value=value)

        # Kopiere und füge die Werte von Spalte J in Spalte P ein
        for row in range(2, end_row + 1):
            value = Quelldatei_Reiter.cell(row=row, column=10).value
            Zieldatei_Reiter.cell(row=row, column=16, value=value)

        # Schreibe den Wert "VES" in jede Zelle der Spalte Q
        for row in range(2, end_row + 1):
            Zieldatei_Reiter.cell(row=row, column=17, value="VES")

        # Schreibe den Wert "11XVE-SALES-PK-P" in jede Zelle der Spalte R
        for row in range(2, end_row + 1):
            Zieldatei_Reiter.cell(row=row, column=18, value="11XVE-SALES-PK-P")

        # Spalte G durchgehen und Zeilen mit "0" löschen
        end_row_zieldatei = Zieldatei_Reiter.max_row  

        for zeile in range(end_row_zieldatei, 1, - 1):
            zelle = Zieldatei_Reiter[f'G{zeile}']
            if zelle.value == 0:
                Zieldatei_Reiter.delete_rows(zeile)

        time.sleep(60)

        # Speichere die Änderungen
        Zieldatei.save(zieldatei_pfad)

        # Die neueste Datei im Quellverzeichnis löschen
        os.remove(quelldatei_pfad)

        print('-------------------------------------------------------------------------------------')
        print('---------die Excel-Datei "VES_SLP_Strom.xlsx" wurde erfolgreich aktualisiert---------')
        print('-------------------------------------------------------------------------------------')
        return True
    
    except Exception as e:

        print(f"Ein Fehler beim Aktualisieren der Datei ""VES_SLP_Strom.xlsx"" ist aufgetreten: {e}")
        return False
    
# Um das Programm auszuführen, soll der Teil unten auskommentiert werden
    
# ausführen()
