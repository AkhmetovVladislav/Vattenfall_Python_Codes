from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from copy import copy
import os
import shutil
import time

def ausführen():
    try:

        # Benutzernamen abrufen
        benutzername = os.getlogin()

        # Quellverzeichnis
        quelle_verzeichnis = r'\\eur.corp.vattenfall.com\vg\Continental\CS\GER\_Data\Sourcing_FP'

        # Zielverzeichnis
        ziel_verzeichnis = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcing - Dokumente\IT\Tools\Vladi\Excel-Ausfüller\source\VREES_Fahrplandaten_Strom'

        # Liste aller Dateien im Quellverzeichnis
        dateien = os.listdir(quelle_verzeichnis)

        # Filtern nach Dateien, die "Fahrplandaten_Strom" im Namen enthalten
        passende_dateien = [datei for datei in dateien if "Fahrplandaten_Strom" in datei]

        # Die neueste Datei auswählen
        neueste_datei = max(passende_dateien, key=lambda x: os.path.getmtime(os.path.join(quelle_verzeichnis, x)))

        # Den vollständigen Pfad zur neuesten Datei erstellen
        quelle_pfad = os.path.join(quelle_verzeichnis, neueste_datei)

        # Ziel-Pfad für die kopierte Datei
        ziel_pfad = os.path.join(ziel_verzeichnis, "Fahrplandaten_Strom.xlsx")

        # Die neueste Datei kopieren und umbenennen
        shutil.copy(quelle_pfad, ziel_pfad)

        # Zieldatei "VREES_Fahrplandaten_Strom.xlsx" mit dem betrachteten Reiter "VREES_Fahrplandaten_Strom"

        zieldatei_pfad = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcingPowerBI - Dokumente\IT\PowerBI_Masterdateien\Excel_Masterdateien\VREES_Fahrplandaten_Strom.xlsx'
        Zieldatei = load_workbook(zieldatei_pfad)
        Zieldatei_Reiter = Zieldatei["VREES_Fahrplandaten_Strom"]

        # Quelldatei "Fahrplandaten_Strom.xlsx" mit dem betrachteten Reiter "Report"

        quelldatei_pfad = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcing - Dokumente\IT\Tools\Vladi\Excel-Ausfüller\source\VREES_Fahrplandaten_Strom\Fahrplandaten_Strom.xlsx'
        Quelldatei = load_workbook(quelldatei_pfad)
        Quelldatei_Reiter = Quelldatei["Report"]

        # Erstelle eine neue Spalte nach der Spalte AI (an der Stelle 36)
        Quelldatei_Reiter.insert_cols(36)
        Quelldatei_Reiter['AJ1'] = "RLM_LITE"

        # Schleife durch jede Zeile, beginnend ab Zeile 2 (Annahme: Überschrift in Zeile 1)
        for row_num in range(2, Quelldatei_Reiter.max_row + 1):
            # Die Zellwerte für AK und AP in der aktuellen Zeile erhalten
            ak_wert = Quelldatei_Reiter[f"AK{row_num}"].value
            ap_wert = Quelldatei_Reiter[f"AP{row_num}"].value
            
            # Überprüfen, ob die Bedingung erfüllt ist und den Wert entsprechend setzen
            if ak_wert == "ZC0" and ap_wert == "Z52":
                Quelldatei_Reiter.cell(row=row_num, column=36, value="X")
            else:
                Quelldatei_Reiter.cell(row=row_num, column=36, value="")

        # Lösche die restlichen Spalten ab Spalte 37
        Quelldatei_Reiter.delete_cols(37,7)

        # Bereinige die Zieldatei
        Zieldatei_Reiter.delete_cols(1,37)

        # Kopieren der ersten 36 Spalten von Quelldatei nach Zieldatei
        for col in range(1, 37):
            quelle_spalte = get_column_letter(col)
            ziel_spalte = get_column_letter(col)
            for row in range(1, Quelldatei_Reiter.max_row + 1):
                zellwert = Quelldatei_Reiter[quelle_spalte + str(row)].value
                Zieldatei_Reiter[ziel_spalte + str(row)] = zellwert

        time.sleep(60)

        # Speichere die Änderungen in der Zieldatei
        Zieldatei.save(zieldatei_pfad)

        # Die neueste Datei im Quellverzeichnis löschen
        os.remove(quelldatei_pfad)

        print('-------------------------------------------------------------------------------------')
        print('---die Excel-Datei "VREES_Fahrplandaten_Strom.xlsx" wurde erfolgreich aktualisiert---')
        print('-------------------------------------------------------------------------------------')
        return True
    
    except Exception as e:

        print(f"Ein Fehler beim Aktualisieren der Datei ""VREES_Fahrplandaten_Strom.xlsx"" ist aufgetreten: {e}")
        return False
    
# Um das Programm auszuführen, soll der Teil unten auskommentiert werden
    
# ausführen()

