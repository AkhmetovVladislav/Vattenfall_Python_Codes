from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from copy import copy
import datetime
import os
import shutil
import xlwings as xw
import time 
import warnings

def ausführen():
    try:

        # Schaltet die Warnungen aus
        warnings.filterwarnings("ignore")

        # Benutzernamen abrufen
        benutzername = os.getlogin()

        # Quellverzeichnis mit den Datum-Ordnern
        quelle_verzeichnis = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcing - Dokumente\IT\PowerBI_Testdateien\Vladislav\Allokationen'

        # Zielverzeichnis
        ziel_verzeichnis = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcing - Dokumente\IT\Tools\Vladi\Excel-Ausfüller\source\Allokation_ANB_Tag_VES_VREES'

        # Durchsuche das Quellverzeichnis nach den Datum-Ordnern
        datum_ordner = [ordner for ordner in os.listdir(quelle_verzeichnis) if os.path.isdir(os.path.join(quelle_verzeichnis, ordner))]

        # Sortiere die Datum-Ordner nach ihrem Namen (also nach Datum)
        datum_ordner.sort(reverse=True)

        # Der neueste Datum-Ordner
        neuester_ordner = datum_ordner[0]

        # Der vollständige Pfad zum neuesten Datum-Ordner
        neuester_ordner_pfad = os.path.join(quelle_verzeichnis, neuester_ordner)

        # Alle Dateien im neuesten Datum-Ordner kopieren
        for datei in os.listdir(neuester_ordner_pfad):
            quelle_pfad = os.path.join(neuester_ordner_pfad, datei)
            ziel_pfad = os.path.join(ziel_verzeichnis, datei)
            shutil.copy(quelle_pfad, ziel_pfad)
                                                            

        # Zieldatei "Allokation_ANB_Tag_VES_VREES.xlsx" mit dem betrachteten Reiter "Allok_Tag_VES"

        zieldatei_pfad = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcingPowerBI - Dokumente\IT\PowerBI_Masterdateien\Excel_Masterdateien\Allokation_ANB_Tag_VES_VREES.xlsx'
        Zieldatei = load_workbook(zieldatei_pfad)
        Zieldatei_Reiter = Zieldatei["Allok_Tag_VES"]

        # 1. Quelldatei "VES-H.xlsx" mit dem betrachteten Reiter "Datenmonitor-Export"

        quelldatei_VES_H_pfad = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcing - Dokumente\IT\Tools\Vladi\Excel-Ausfüller\source\Allokation_ANB_Tag_VES_VREES\VES-H.xlsx'
        Quelldatei_VES_H = load_workbook(quelldatei_VES_H_pfad)
        Quelldatei_VES_H_Reiter = Quelldatei_VES_H["Datenmonitor-Export"]

        # 2. Quelldatei "VES-L.xlsx" mit dem betrachteten Reiter "Datenmonitor-Export"

        quelldatei_VES_L_pfad = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcing - Dokumente\IT\Tools\Vladi\Excel-Ausfüller\source\Allokation_ANB_Tag_VES_VREES\VES-L.xlsx'
        Quelldatei_VES_L = load_workbook(quelldatei_VES_L_pfad)
        Quelldatei_VES_L_Reiter = Quelldatei_VES_L["Datenmonitor-Export"]

        # 3. Quelldatei "VREES-H.xlsx" mit dem betrachteten Reiter "Datenmonitor-Export"

        quelldatei_VREES_H_pfad = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcing - Dokumente\IT\Tools\Vladi\Excel-Ausfüller\source\Allokation_ANB_Tag_VES_VREES\VREES-H.xlsx'
        Quelldatei_VREES_H = load_workbook(quelldatei_VREES_H_pfad)
        Quelldatei_VREES_H_Reiter = Quelldatei_VREES_H["Datenmonitor-Export"]

        # 4. Quelldatei "VREES-L.xlsx" mit dem betrachteten Reiter "Datenmonitor-Export"

        quelldatei_VREES_L_pfad = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcing - Dokumente\IT\Tools\Vladi\Excel-Ausfüller\source\Allokation_ANB_Tag_VES_VREES\VREES-L.xlsx'
        Quelldatei_VREES_L = load_workbook(quelldatei_VREES_L_pfad)
        Quelldatei_VREES_L_Reiter = Quelldatei_VREES_L["Datenmonitor-Export"]

        # Übersetzungstabelle "Allokation_ANB_Tag_VES_VREES_Übersetzer_ANB.xlsx" mit dem betrachteten Reiter "DVGW-ANB"

        uebersetzungstabelle_pfad = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcingPowerBI - Dokumente\IT\PowerBI_Masterdateien\Excel_Masterdateien\Allokation_ANB_Tag_VES_VREES_Übersetzer_ANB.xlsx'
        uebersetzungstabelle = load_workbook(uebersetzungstabelle_pfad)
        uebersetzungstabelle_Reiter = uebersetzungstabelle["DVGW-ANB"]

        # 1.Schritt
        # Kopieren der Daten von VES-L.xlsx in VES-H.xlsx mit Formatierung
        max_row_h = Quelldatei_VES_H_Reiter.max_row
        max_column_h = Quelldatei_VES_H_Reiter.max_column
        max_row_l = Quelldatei_VES_L_Reiter.max_row

        for row in range(2, max_row_l + 1):  # Ab der zweiten Zeile von VES-L.xlsx
            for column in range(1, max_column_h + 1):  # Von Spalte A bis Spalte AM in VES-H.xlsx
                cell_value = Quelldatei_VES_L_Reiter.cell(row=row, column=column).value
                source_cell = Quelldatei_VES_L_Reiter.cell(row=row, column=column)
                target_cell = Quelldatei_VES_H_Reiter.cell(row=max_row_h + row - 1, column=column)
                target_cell.value = cell_value
                
                # Übernahme der Formatierung von der Quelldatei (VES-L.xlsx)
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

        # 2.Schritt
        # Überprüfe und lösche Zeilen in VES-H.xlsx, wenn in Spalte E der Wert "RLMMT" steht
        max_row_h = Quelldatei_VES_H_Reiter.max_row

        for row in range(max_row_h, 1, -1):  # Durchlaufe die Zeilen rückwärts
            cell_value = Quelldatei_VES_H_Reiter.cell(row=row, column=5).value  # Spalte E entspricht der 5. Spalte
            if cell_value == "RLMMT":
                # Wenn "RLMMT" in Spalte E steht, lösche die gesamte Zeile
                Quelldatei_VES_H_Reiter.delete_rows(row)

        # 3.Schritt
        # Kopieren der Daten von VREES-H.xlsx in VES-H.xlsx mit Formatierung
        max_row_h = Quelldatei_VES_H_Reiter.max_row
        max_row_h_vrees = Quelldatei_VREES_H_Reiter.max_row

        for row_vrees in range(2, max_row_h_vrees + 1):  # Ab der zweiten Zeile von VREES-H.xlsx
            for column_vrees in range(1, max_column_h + 1):  # Von Spalte A bis Spalte AM in VES-H.xlsx
                cell_value_vrees = Quelldatei_VREES_H_Reiter.cell(row=row_vrees, column=column_vrees).value
                source_cell_vrees = Quelldatei_VREES_H_Reiter.cell(row=row_vrees, column=column_vrees)
                target_cell_vrees = Quelldatei_VES_H_Reiter.cell(row=max_row_h + row_vrees - 1, column=column_vrees)
                target_cell_vrees.value = cell_value_vrees
                
                # Übernahme der Formatierung von der Quelldatei (VREES-H.xlsx)
                target_cell_vrees.font = copy(source_cell_vrees.font)
                target_cell_vrees.border = copy(source_cell_vrees.border)
                target_cell_vrees.fill = copy(source_cell_vrees.fill)
                target_cell_vrees.number_format = copy(source_cell_vrees.number_format)
                target_cell_vrees.protection = copy(source_cell_vrees.protection)
                target_cell_vrees.alignment = copy(source_cell_vrees.alignment)

        # 4.Schritt
        # Kopieren der Daten von VREES-L.xlsx in VES-H.xlsx mit Formatierung
        max_row_h = Quelldatei_VES_H_Reiter.max_row
        max_row_l_vrees = Quelldatei_VREES_L_Reiter.max_row

        for row_vrees in range(2, max_row_l_vrees + 1):  # Ab der zweiten Zeile von VREES-L.xlsx
            for column_vrees in range(1, max_column_h + 1):  # Von Spalte A bis Spalte AM in VES-H.xlsx
                cell_value_vrees = Quelldatei_VREES_L_Reiter.cell(row=row_vrees, column=column_vrees).value
                source_cell_vrees = Quelldatei_VREES_L_Reiter.cell(row=row_vrees, column=column_vrees)
                target_cell_vrees = Quelldatei_VES_H_Reiter.cell(row=max_row_h + row_vrees - 1, column=column_vrees)
                target_cell_vrees.value = cell_value_vrees
                
                # Übernahme der Formatierung von der Quelldatei (VREES-L.xlsx)
                target_cell_vrees.font = copy(source_cell_vrees.font)
                target_cell_vrees.border = copy(source_cell_vrees.border)
                target_cell_vrees.fill = copy(source_cell_vrees.fill)
                target_cell_vrees.number_format = copy(source_cell_vrees.number_format)
                target_cell_vrees.protection = copy(source_cell_vrees.protection)
                target_cell_vrees.alignment = copy(source_cell_vrees.alignment)

        # Zwischenschritt
        # Durchlaufe die Zeilen in Spalte D von VES-H.xlsx
        max_row_h = Quelldatei_VES_H_Reiter.max_row

        for row_h in range(2, max_row_h + 1):
            cell_value_d = Quelldatei_VES_H_Reiter.cell(row=row_h, column=4).value  # Spalte D entspricht der 4. Spalte
            if cell_value_d is None or str(cell_value_d).strip() == "":
                # Wenn die Zelle leer ist, setze den Wert auf "4043581000034"
                Quelldatei_VES_H_Reiter.cell(row=row_h, column=4).value = "4043581000034"

        # 5.Schritt
        # Durchlaufe die Zeilen in Spalte E von VES-H.xlsx
        max_row_h = Quelldatei_VES_H_Reiter.max_row

        for row_h in range(2, max_row_h + 1):
            cell_value_e = Quelldatei_VES_H_Reiter.cell(row=row_h, column=5).value  # Spalte E entspricht der 5. Spalte
            if cell_value_e == "SLPANA" or cell_value_e == "SLPSYN":
                # Wenn "SLPANA" oder "SLPSYN" in Spalte E steht, schreibe "kein Brennwert" in Spalte F
                Quelldatei_VES_H_Reiter.cell(row=row_h, column=6, value="kein Brennwert")

        # 6.Schritt
        # Durchlaufe die Zeilen in Spalte D von VES-H.xlsx
                
        for row_h in range(2, max_row_h + 1):
            cell_value_d = Quelldatei_VES_H_Reiter.cell(row=row_h, column=4).value  # Spalte D entspricht der 4. Spalte
            if isinstance(cell_value_d, str) and cell_value_d.strip().replace('.', '', 1).isdigit():
                # Überprüfen, ob der Zellenwert eine Zahl als Text ist
                Quelldatei_VES_H_Reiter.cell(row=row_h, column=4).value = float(cell_value_d)

        # 7.Schritt
        # Bestimme die letzte ausgefüllte Spalte der 2. Zeile
        max_column_h = Quelldatei_VES_H_Reiter.max_column
        last_filled_column = None

        for column_h in range(1, max_column_h + 2):
            cell_value = Quelldatei_VES_H_Reiter.cell(row=2, column=column_h).value
            if cell_value == '' or cell_value is None:
                last_filled_column = column_h - 1
                break

        # Durchlaufe die Zeilen und überprüfe die Bedingungen
        max_row_h = Quelldatei_VES_H_Reiter.max_row

        for row_h in range(2, max_row_h + 1):
            cell_value_b = Quelldatei_VES_H_Reiter.cell(row=row_h, column=2).value  # Spalte B entspricht der 2. Spalte
            cell_value_e = Quelldatei_VES_H_Reiter.cell(row=row_h, column=5).value  # Spalte E entspricht der 5. Spalte
            if cell_value_b in ["THE0BFH008470000", "THE0BFL008480000"] and cell_value_e in ["SLPANA", "SLPSYN"]:
                # Wenn die Bedingungen erfüllt sind, lösche die Daten von den letzten 2 ausgefüllten Spalten
                for column_h in range(last_filled_column + 1, last_filled_column + 3):
                    Quelldatei_VES_H_Reiter.cell(row=row_h, column=column_h).value = None

        # 8.Schritt
        # Durchlaufe die Zellen in Spalte H von VES-H.xlsx
        max_row_h = Quelldatei_VES_H_Reiter.max_row

        for row_h in range(2, max_row_h + 1):
            summe = 0
            for column_h in range(9, max_column_h + 1):  # Durchlaufe die Spalten von I bis AM
                cell_value = Quelldatei_VES_H_Reiter.cell(row=row_h, column=column_h).value
                if isinstance(cell_value, (int, float)):
                    summe += cell_value

            # Trage die Summe in die Zelle in Spalte H ein
            Quelldatei_VES_H_Reiter.cell(row=row_h, column=8, value=summe)

        # 9.Schritt
        # Durchlaufe die Zellen von Spalte H bis Spalte AM von VES-H.xlsx
        max_row_h = Quelldatei_VES_H_Reiter.max_row
        max_column_h = Quelldatei_VES_H_Reiter.max_column

        for row_h in range(2, max_row_h + 1):
            for column_h in range(8, max_column_h + 1):  # Durchlaufe die Spalten von H bis AM
                cell = Quelldatei_VES_H_Reiter.cell(row=row_h, column=column_h)
                cell_value = cell.value
                if isinstance(cell_value, (int, float)):
                    # Teile den Wert durch 1000
                    new_value = cell_value / 1000
                    # Speichere die ursprüngliche Formatierung
                    formatierung = cell.number_format
                    # Schreibe den neuen Wert in die Zelle mit der gleichen Formatierung
                    cell.value = new_value
                    cell.number_format = formatierung

        # 10.Schritt
        # Spalte C bereinigen
        for row_h in range(2, max_row_h + 1):
            Quelldatei_VES_H_Reiter.cell(row=row_h, column=3).value = None

        # Durchlaufe die Zellen in Spalte C von VES-H.xlsx
        max_row_h = Quelldatei_VES_H_Reiter.max_row

        for row_h in range(2, max_row_h + 1):
            # Wert in Spalte D von VES-H.xlsx in der aktuellen Zeile erhalten
            value_d_ves_h = Quelldatei_VES_H_Reiter.cell(row=row_h, column=4).value
            
            # Durchlaufe die Werte in Spalte A der Übersetzungstabelle
            max_row_uebersetzungstabelle = uebersetzungstabelle_Reiter.max_row
            for row_uebersetzungstabelle in range(2, max_row_uebersetzungstabelle + 1):
                # Wert in Spalte A der Übersetzungstabelle in der aktuellen Zeile erhalten
                value_a_uebersetzungstabelle = uebersetzungstabelle_Reiter.cell(row=row_uebersetzungstabelle, column=1).value
                
                # Wenn der Wert in Spalte D von VES-H.xlsx mit einem Wert in Spalte A der Übersetzungstabelle übereinstimmt
                if value_d_ves_h == value_a_uebersetzungstabelle:
                    # Wert in Spalte B der Übersetzungstabelle in die Zelle in Spalte C von VES-H.xlsx einfügen
                    value_b_uebersetzungstabelle = uebersetzungstabelle_Reiter.cell(row=row_uebersetzungstabelle, column=2).value
                    Quelldatei_VES_H_Reiter.cell(row=row_h, column=3).value = value_b_uebersetzungstabelle
                    break  # Sobald ein übereinstimmender Wert gefunden wurde, die Schleife beenden
            
            else:  # Wenn keine Übereinstimmung gefunden wurde
                # "kein Wert" in die Zelle in Spalte C von VES-H.xlsx einfügen
                Quelldatei_VES_H_Reiter.cell(row=row_h, column=3).value = "kein Wert"

        # 11.Schritt 
        # die Werten in die Zieldatei übertragen

        # Wert der Zelle A2 aus "VES-H.xlsx" kopieren
        wert_a2_ves_h = Quelldatei_VES_H_Reiter['A2'].value

        # Durchlaufe die Zieldatei "Allokation_ANB_Tag_VES_VREES_kopie.xlsx" in der Spalte A (von A2 bis zum Ende)
        max_row_zieldatei = Zieldatei_Reiter.max_row

        found = False  # Eine Variable, um den Status des Fundes zu verfolgen

        for row_zieldatei in range(2, max_row_zieldatei + 1):
            zieldatei_zelle_value = Zieldatei_Reiter.cell(row=row_zieldatei, column=1).value  # Wert in Spalte A der Zieldatei
            if zieldatei_zelle_value == wert_a2_ves_h:
                found = True

                # Lösche eine bestimmte Anzahl von Zeilen von Spalte A bis AM für die gefundenen Zeilen
                anz_zeilen_loeschen = Quelldatei_VES_H_Reiter.max_row - 1  # Anzahl der Zeilen in "VES-H.xlsx" außer der ersten Zeile
                for row in range(row_zieldatei, row_zieldatei + anz_zeilen_loeschen):
                    for column in range(1, 40):  # Annahme: Es gibt insgesamt 40 Spalten von A bis AM
                        Zieldatei_Reiter.cell(row=row, column=column).value = ""

                # Kopiere die Werte von "VES-H.xlsx" in die Zieldatei und übernehme die Formatierung
                for row_quelldatei in range(2, Quelldatei_VES_H_Reiter.max_row + 1):
                    for column_quelldatei in range(1, 40):  # Annahme: Es gibt insgesamt 40 Spalten von A bis AM
                        cell_value = Quelldatei_VES_H_Reiter.cell(row=row_quelldatei, column=column_quelldatei).value
                        source_cell = Quelldatei_VES_H_Reiter.cell(row=row_quelldatei, column=column_quelldatei)
                        target_cell = Zieldatei_Reiter.cell(row=row_zieldatei + row_quelldatei - 2, column=column_quelldatei)
                        target_cell.value = cell_value
                        
                        # Übernahme der Formatierung von der Quelldatei (VES-H.xlsx)
                        target_cell.font = copy(source_cell.font)
                        target_cell.border = copy(source_cell.border)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.number_format = copy(source_cell.number_format)
                        target_cell.protection = copy(source_cell.protection)
                        target_cell.alignment = copy(source_cell.alignment)

                break

        if not found:
            # Einfügen der Werte und Formatierungen von "VES-H.xlsx" am Ende der Zieldatei
            for row_quelldatei in range(2, Quelldatei_VES_H_Reiter.max_row + 1):
                for column_quelldatei in range(1, 40):  # Annahme: Es gibt insgesamt 40 Spalten von A bis AM
                    cell_value = Quelldatei_VES_H_Reiter.cell(row=row_quelldatei, column=column_quelldatei).value
                    source_cell = Quelldatei_VES_H_Reiter.cell(row=row_quelldatei, column=column_quelldatei)
                    target_cell = Zieldatei_Reiter.cell(row=max_row_zieldatei + row_quelldatei - 1, column=column_quelldatei)
                    target_cell.value = cell_value
                    
                    # Übernahme der Formatierung von der Quelldatei (VES-H.xlsx)
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = copy(source_cell.number_format)
                    target_cell.protection = copy(source_cell.protection)
                    target_cell.alignment = copy(source_cell.alignment)

        # Speichern der Änderungen in VES-H.xlsx
        Zieldatei.save(zieldatei_pfad)

        os.remove(quelldatei_VES_H_pfad)
        os.remove(quelldatei_VES_L_pfad)
        os.remove(quelldatei_VREES_H_pfad)
        os.remove(quelldatei_VREES_L_pfad)

        time.sleep(60)

        # Benutzernamen erhalten
        username = os.getlogin()

        # Pfad zur Zieldatei mit allgemeinem Format
        makro_pfad = fr"C:\Users\{username}\Vattenfall AB\CSGERSourcing - Dokumente\IT\Tools\Vladi\Excel-Ausfüller\source\Allokation_ANB_Tag_VES_VREES\makro_allokation.xlsm"

        # Öffnen der Zieldatei im Hintergrund
        app = xw.App(visible=False)
        Makro_Datei = xw.Book(makro_pfad)

        # Auswahl des Reiters
        Makro_Datei_Reiter = Makro_Datei.sheets["makro_allokation"]

        # Ausführen des Makros "Zahlen_zu_Datum"
        Makro_Datei.macro("Makro_allokation")()

        # Speichern und Schließen der Datei
        Makro_Datei.save()
        Makro_Datei.close()

        # Beenden der Excel-Anwendung
        app.quit()

        # Ausgabe "Erfolgreich"

        print('-------------------------------------------------------------------------------------')
        print('--die Excel-Datei "Allokation_ANB_Tag_VES_VREES.xlsx" wurde erfolgreich aktualisiert-')
        print('-------------------------------------------------------------------------------------')

        return True
    
    except Exception as e:

        print(f'Ein Fehler beim Aktualisieren der Datei "Allokation_ANB_Tag_VES_VREES.xlsx" ist aufgetreten: {e}')
        return False
    
# Um das Programm auszuführen, soll der Teil unten auskommentiert werden
    
# ausführen()
