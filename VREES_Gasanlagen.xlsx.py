from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os
import shutil
import time

def ausführen():
    try:

        # Benutzernamen abrufen
        benutzername = os.getlogin()

        # Quellverzeichnis
        quelle_verzeichnis = r'\\eur.corp.vattenfall.com\v\_data\Sales_SFTP\Regiocom\In\sourcing_gas_report'

        # Zieldateiverzeichnis erstellen
        ziel_verzeichnis = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcing - Dokumente\IT\Tools\Vladi\Excel-Ausfüller\source\VREES_Gasanlagen'

        # Liste aller Dateien im Quellverzeichnis
        dateien = os.listdir(quelle_verzeichnis)

        # Filtern nach Dateien, die "Fahrplandaten_Gas" im Namen enthalten
        passende_dateien = [datei for datei in dateien if "Fahrplandaten_Gas" in datei]

        # Die neueste Datei auswählen
        neueste_datei = max(passende_dateien, key=lambda x: os.path.getmtime(os.path.join(quelle_verzeichnis, x)))

        # Den vollständigen Pfad zur neuesten Datei erstellen
        quelle_pfad = os.path.join(quelle_verzeichnis, neueste_datei)

        # Ziel-Pfad für die kopierte Datei
        ziel_pfad = os.path.join(ziel_verzeichnis, "Fahrplandaten_Gas.xlsx")

        # Die neueste Datei kopieren und umbenennen
        shutil.copy(quelle_pfad, ziel_pfad)

        #Zieldatei "VREES_Gasanlagen.xlsx" mit dem betrachteten Reiter "Fahrplandaten_Gas_VREES"

        zieldatei_pfad = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcingPowerBI - Dokumente\IT\PowerBI_Masterdateien\Excel_Masterdateien\VREES_Gasanlagen.xlsx'
        Zieldatei = load_workbook(zieldatei_pfad)
        Zieldatei_Reiter = Zieldatei["Fahrplandaten_Gas_VREES"]

        #Quelldatei "Fahrplandaten_Gas.xlsx" mit dem betrachteten Reiter "Report"

        quelldatei_pfad = rf'C:\Users\{benutzername}\Vattenfall AB\CSGERSourcing - Dokumente\IT\Tools\Vladi\Excel-Ausfüller\source\VREES_Gasanlagen\Fahrplandaten_Gas.xlsx'
        Quelldatei = load_workbook(quelldatei_pfad)
        Quelldatei_Reiter = Quelldatei["Report"]

        # Lösche die ersten 19 Spalten
        Zieldatei_Reiter.delete_cols(1,19)

        # Schreibe Überschriften neu
        ueberschriften = [
            "VERBRAUCHSTELLE_PLZ", "VERBRAUCHSTELLE_ORT", "VERBRAUCHSTELLE_MALOID",
            "VERSORGUNGS_BEGINN", "VERSORGUNGS_ENDE", "GUELTIG_AB", "EIC_BILANZKREIS",
            "VNB_NAME", "JAHRESARBEIT", "JAHRESARBEIT_AUS_BESTELLUNG",
            "ZAEHLVERFAHREN", "Zeitreihentyp", "JVP Best Of"
        ]

        for idx, ueberschrift in enumerate(ueberschriften, start=1):
            Zieldatei_Reiter[f"{get_column_letter(idx)}1"] = ueberschrift

        # 1.Schritt: die unnötigen Spalten löschen    
        Quelldatei_Reiter.delete_cols(28,15)
        Quelldatei_Reiter.delete_cols(24,3)
        Quelldatei_Reiter.delete_cols(21,2)
        Quelldatei_Reiter.delete_cols(16,4)
        Quelldatei_Reiter.delete_cols(12,3)
        Quelldatei_Reiter.delete_cols(1,4)

        # 2.Schritt: Kopiere die übrigen Spalten ab 2.Zeile
        for row in range(2, Quelldatei_Reiter.max_row + 1):  # Starte von der zweiten Zeile bis zu letzter Zeile
            for col in range(1, 12):  #die ersten 12 Spalten
                Zieldatei_Reiter.cell(row=row, column=col).value = Quelldatei_Reiter.cell(row=row, column=col).value

        # 2.1 Schritt (zusätzlicher), falls in Spalte K nichts steht

        for cell in Zieldatei_Reiter['K']:
            if cell.value is None or cell.value == "":
                # Wenn die Zelle leer ist, setzen Sie den Wert entsprechend der Bedingung in Spalte J
                row_number = cell.row
                value_j = Zieldatei_Reiter.cell(row=row_number, column=10).value
                if value_j is not None and value_j != "" and int(value_j) < 1500000:
                    cell.value = "E02"
                else:
                    cell.value = "E01"


        # 3.Schritt: Schreibe die Formel in Spalte L ab der zweiten Zeile
        for row in range(2, Zieldatei_Reiter.max_row + 1):
            if Zieldatei_Reiter[f"K{row}"].value in ("E02", None):
                Zieldatei_Reiter[f"L{row}"] = "SLS"
            else:
                Zieldatei_Reiter[f"L{row}"] = "LGS"

        # 4.Schritt: Schreibe die Formel in Spalte M ab der zweiten Zeile
        for row in range(2, Zieldatei_Reiter.max_row + 1):
            if Zieldatei_Reiter[f"I{row}"].value == "":
                Zieldatei_Reiter[f"M{row}"] = Zieldatei_Reiter[f"J{row}"].value
            else:
                Zieldatei_Reiter[f"M{row}"] = Zieldatei_Reiter[f"I{row}"].value

        time.sleep(60)

        # Speichere die Änderungen in der Zieldatei
        Zieldatei.save(zieldatei_pfad)

        # Die neueste Datei im Quellverzeichnis löschen
        os.remove(quelldatei_pfad)

        print('-------------------------------------------------------------------------------------')
        print('--------die Excel-Datei "VREES_Gasanlagen.xlsx" wurde erfolgreich aktualisiert-------')
        print('-------------------------------------------------------------------------------------')
        return True
    
    except Exception as e:

        print(f"Ein Fehler beim Aktualisieren der Datei ""VREES_Gasanlagen.xlsx"" ist aufgetreten: {e}")
        return False
    
# Um das Programm auszuführen, soll der Teil unten auskommentiert werden
    
# ausführen()
